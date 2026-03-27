from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash, logout
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from django.db.models.functions import TruncMonth, TruncDate
from django.db.models import Count, Avg, Sum, F
from django.db.models import Q

from .decorators import role_required
from .forms import OwnerProfileUpdateForm , EventProfileUpdateForm

from events.models import EventRegistration
from business.models import business as BusinessModel, Review, Inquiry, ViewTracking

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
import json
from datetime import datetime



@login_required
@role_required(allowed_roles=["event_organizer", "owner"])
def export_attendees(request):
    # Fetch data similar to attendeesevent
    registrations = EventRegistration.objects.filter(event__organizer=request.user).select_related('event', 'user')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendees"

    # Headers
    headers = ['Name', 'Email', 'Event', 'Status', 'RSVP Date']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    row_num = 2
    status_mapping = {
        'registered': 'Pending',
        'attended': 'Confirmed',
        'cancelled': 'Cancelled'
    }

    for reg in registrations:
        name = f"{reg.user.first_name} {reg.user.last_name}".strip() or reg.user.username
        email = reg.user.email
        event_title = reg.event.title
        status = status_mapping.get(reg.status, reg.status)
        rsvp_date = reg.registration_date.date().strftime('%Y-%m-%d')

        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=2, value=email)
        ws.cell(row=row_num, column=3, value=event_title)
        ws.cell(row=row_num, column=4, value=status)
        ws.cell(row=row_num, column=5, value=rsvp_date)

        row_num += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendees.xlsx'

    # Save to response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())

    return response


# ================= DASHBOARDS =================

@login_required
@role_required(allowed_roles=["owner"])
def business_owner_dashboard(request):
    from business.models import business as BusinessModel, Review, Inquiry

    owner_businesses = BusinessModel.objects.filter(owner_id=request.user)

    total_views = owner_businesses.aggregate(total=Sum('views_count'))['total'] or 0
    active_listings = owner_businesses.filter(is_active=True).count()

    avg_rating = Review.objects.filter(business__owner_id=request.user).aggregate(avg=Avg('rating'))['avg'] or 0
    avg_rating = round(avg_rating, 2)

    inquiries_count = Inquiry.objects.filter(business__owner_id=request.user).count()

    recent_reviews = Review.objects.filter(business__owner_id=request.user).order_by('-created_at')[:5]

    return render(request, "localcommunity/businessowner/businessowner.html", {
        'total_views': total_views,
        'active_listings': active_listings,
        'avg_rating': avg_rating,
        'inquiries_count': inquiries_count,
        'recent_reviews': recent_reviews,
    })


@login_required
@role_required(allowed_roles=["user"])
def user_dashboard(request):
    return render(request, "core/index.html")


# ================= BUSINESS OWNER PAGES =================

def businessownerdashboard(request):
    return render(request, "localcommunity/businessowner/businessowner.html")


@login_required
@role_required(allowed_roles=["owner"])
def mybusinesses(request):
    user_businesses = BusinessModel.objects.filter(owner_id=request.user)

    annotated_businesses = user_businesses.annotate(
        avg_rating=Avg('reviews__rating'),
        total_reviews=Count('reviews')
    )

    businesses = []
    for b in annotated_businesses:
        businesses.append({
            'id': b.id,
            'name': b.business_name,
            'category': b.category or 'N/A',
            'address': b.address,
            'image_url': b.image or 'https://via.placeholder.com/380x220?text=No+Image',
            'status': 'Active' if b.is_active else 'Pending',
            'avg_rating': round(b.avg_rating or 0, 1) if b.total_reviews > 0 else 'N/A',
            'total_reviews': b.total_reviews,
            'views_count': b.views_count,
        })

    return render(request, "localcommunity/businessowner/mybusinesses.html", {
        'businesses': businesses
    })


@login_required
@role_required(allowed_roles=["owner"])
def edit_business(request, business_id):
    # Placeholder: hook this to the actual edit form and process updates.
    business_obj = BusinessModel.objects.filter(id=business_id, owner_id=request.user).first()
    if not business_obj:
        return redirect('localcommunity:mybusinesses')

    # laod edit page or process post data in production
    return render(request, 'localcommunity/businessowner/editbusiness.html', {'business': business_obj})


@login_required
@role_required(allowed_roles=["owner"])
def delete_business(request, business_id):
    business_obj = BusinessModel.objects.filter(id=business_id, owner_id=request.user).first()
    if business_obj:
        business_obj.delete()
    return redirect('localcommunity:mybusinesses')


def addbusiness(request):
    return render(request, "localcommunity/businessowner/addbusiness.html")


@login_required
@role_required(allowed_roles=["owner"])
def reviews(request):
    filter_type = request.GET.get('filter', 'all')

    # Reviews for businesses owned by current user
    business_qs = BusinessModel.objects.filter(owner_id=request.user)
    reviews_qs = Review.objects.filter(business__in=business_qs).select_related('business', 'user')

    if filter_type == 'critical':
        reviews_qs = reviews_qs.filter(rating__lte=3)

    total_reviews = reviews_qs.count()
    avg_rating = reviews_qs.aggregate(avg=Avg('rating'))['avg'] or 0
    avg_rating = round(avg_rating, 1) if total_reviews > 0 else 0

    now = timezone.now()
    reviews_this_month = reviews_qs.filter(created_at__year=now.year, created_at__month=now.month).count()
    unreplied_count = reviews_qs.filter(is_replied=False).count()

    recent_reviews = reviews_qs.order_by('-created_at')[:5]

    return render(request, "localcommunity/businessowner/reviews.html", {
        'total_reviews': total_reviews,
        'avg_rating': avg_rating,
        'reviews_this_month': reviews_this_month,
        'unreplied_count': unreplied_count,
        'recent_reviews': recent_reviews,
        'filter_type': filter_type,
    })


@login_required
@role_required(allowed_roles=["owner"])
def analyticsbusiness(request):
    """
    Analytics Dashboard with time range filtering and views over time graph
    """
    
    # Get logged-in owner's businesses
    owner_businesses = BusinessModel.objects.filter(owner_id=request.user)
    
    # Get time range from query params (default: 30 days)
    selected_range = int(request.GET.get('range', 30))
    if selected_range not in [7, 30, 90]:
        selected_range = 30
    
    # Calculate date filter
    now = timezone.now()
    start_date = now - timedelta(days=selected_range)
    
    # ============ AGGREGATE METRICS ============
    
    # Total Views across all owner's businesses
    total_views = ViewTracking.objects.filter(
        business__in=owner_businesses,
        created_at__gte=start_date
    ).count()
    
    # Unique Visitors (distinct users or sessions)
    unique_visitors = ViewTracking.objects.filter(
        business__in=owner_businesses,
        created_at__gte=start_date
    ).values('user_id', 'session_id').distinct().count()
    
    # Total Reviews
    total_reviews = Review.objects.filter(
        business__in=owner_businesses,
        created_at__gte=start_date
    ).count()
    
    # Total Inquiries
    total_inquiries = Inquiry.objects.filter(
        business__in=owner_businesses,
        created_at__gte=start_date
    ).count()
    
    # ============ VIEWS OVER TIME GRAPH DATA (MONTHLY) ============
    
    # Group views by month and get counts
    views_by_month = ViewTracking.objects.filter(
        business__in=owner_businesses,
        created_at__gte=start_date
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    # Format chart data - ALL 12 MONTHS OF YEAR
    chart_labels = []
    chart_values = []
    
    # Create dictionary of views by month
    views_dict = {item['month'].date(): item['count'] for item in views_by_month}
    
    # Get current year
    current_year = now.year
    
    # Generate all 12 months for the year
    for month_num in range(1, 13):
        # Format label - "Jan", "Feb", "Mar", etc.
        label = datetime(current_year, month_num, 1).strftime('%b')
        chart_labels.append(label)
        
        # Get view count for this month or 0
        month_date = datetime(current_year, month_num, 1).date()
        chart_values.append(views_dict.get(month_date, 0))
    
    # ============ GROWTH METRICS (SIMPLE) ============
    # For growth percentages, calculate previous period
    prev_start = start_date - timedelta(days=selected_range)
    
    prev_views = ViewTracking.objects.filter(
        business__in=owner_businesses,
        created_at__gte=prev_start,
        created_at__lt=start_date
    ).count()
    
    prev_reviews = Review.objects.filter(
        business__in=owner_businesses,
        created_at__gte=prev_start,
        created_at__lt=start_date
    ).count()
    
    prev_inquiries = Inquiry.objects.filter(
        business__in=owner_businesses,
        created_at__gte=prev_start,
        created_at__lt=start_date
    ).count()
    
    # Calculate growth percentages
    views_growth = calculate_growth(prev_views, total_views)
    reviews_growth = calculate_growth(prev_reviews, total_reviews)
    inquiries_growth = calculate_growth(prev_inquiries, total_inquiries)
    
    context = {
        'total_views': total_views,
        'unique_visitors': unique_visitors,
        'total_reviews': total_reviews,
        'total_inquiries': total_inquiries,
        'chart_labels': json.dumps(chart_labels),
        'chart_values': json.dumps(chart_values),
        'selected_range': selected_range,
        'views_growth': views_growth,
        'reviews_growth': reviews_growth,
        'inquiries_growth': inquiries_growth,
    }
    
    return render(request, "localcommunity/businessowner/analyticsbusiness.html", context)


def calculate_growth(prev_value, current_value):
    """Helper to calculate percentage growth"""
    if prev_value == 0:
        return 0 if current_value == 0 else 100
    growth = ((current_value - prev_value) / prev_value) * 100
    return round(growth, 1)



# ================= SETTINGS (FIXED) =================

@login_required
@role_required(allowed_roles=["owner"])
def settingsbusiness(request):

    user = request.user

    if request.method == 'POST':
        form = OwnerProfileUpdateForm(request.POST, instance=user)

        if form.is_valid():

            # ✅ DON'T SAVE YET
            user_obj = form.save(commit=False)

            # password fields
            current_password = form.cleaned_data.get('current_password')
            new_password = form.cleaned_data.get('new_password')
            confirm_password = form.cleaned_data.get('confirm_password')

            # 👉 if user trying to change password
            if current_password or new_password or confirm_password:

                if not user.check_password(current_password):
                    messages.error(request, "Current password is incorrect")
                    return render(request, 'localcommunity/businessowner/settingsbusiness.html', {"form": form})

                if new_password != confirm_password:
                    messages.error(request, "New passwords do not match")
                    return render(request, 'localcommunity/businessowner/settingsbusiness.html', {"form": form})

                # ✅ set new password
                user_obj.set_password(new_password)

                # save BEFORE updating session
                user_obj.save()

                # keep user logged in
                update_session_auth_hash(request, user_obj)

                messages.success(request, "Password updated successfully")

            else:
                # ✅ save normal profile updates
                user_obj.save() 
                messages.success(request, "Profile updated successfully")

            return redirect('localcommunity:businessownerdashboard')

        else:
            print(form.errors)

    else:
        form = OwnerProfileUpdateForm(instance=user)

    return render(request, 'localcommunity/businessowner/settingsbusiness.html', {"form": form})


# ================= LOGOUT =================

def logoutbusiness(request):
    logout(request)
    return redirect('login')  # change to your login URL name


# ================= EVENT STUDIO =================

from django.utils import timezone
from events.models import Event, EventRegistration

@login_required
@role_required(allowed_roles=["event_organizer"])
def eventstudio(request):
    if not request.user.is_authenticated:
        return redirect('login')

    today = timezone.localdate()
    user_events = Event.objects.filter(organizer=request.user)

    upcoming_events = user_events.filter(event_date__gte=today).order_by('event_date')

    total_events = user_events.count()
    total_rsvps = EventRegistration.objects.filter(event__organizer=request.user).count()
    attended = EventRegistration.objects.filter(event__organizer=request.user, status='attended').count()
    avg_attendance = round((attended / total_rsvps * 100), 2) if total_rsvps else 0

    event_cards = []
    for ev in upcoming_events:
        rsvp_count = EventRegistration.objects.filter(event=ev).count()
        max_participants = ev.max_participants or 1
        progress_percent = min(100, int((rsvp_count / max_participants) * 100)) if max_participants > 0 else 0
        event_cards.append({
            'title': ev.title,
            'approval_status': ev.approval_status,
            'event_date': ev.event_date,
            'start_time': ev.start_time,
            'rsvp_count': rsvp_count,
            'max_participants': max_participants,
            'progress_percent': progress_percent,
            'status_display': 'Published' if ev.approval_status == 'approved' else ('Pending' if ev.approval_status == 'pending' else 'Cancelled'),
        })

    return render(request, "localcommunity/eventstudio/eventstudio.html", {
        'total_events': total_events,
        'total_rsvps': total_rsvps,
        'avg_attendance': avg_attendance,
        'total_comments': 0,
        'upcoming_events': event_cards,
    })

@login_required
@role_required(allowed_roles=["event_organizer"])
def eventstudiodashboard(request):
    return render(request, "localcommunity/eventstudio/eventstudio.html")


def myevents(request):
    return render(request, "localcommunity/eventstudio/myevents.html")


def createevent(request):
    return render(request, "localcommunity/eventstudio/createevent.html")


@login_required
@role_required(allowed_roles=["event_organizer", "owner"])
def attendeesevent(request):
    # Fetch all registrations for events organized by the user
    registrations = EventRegistration.objects.filter(event__organizer=request.user).select_related('event', 'user')

    # Prepare attendee data
    attendees = []
    for reg in registrations:
        status_mapping = {
            'registered': 'Pending',
            'attended': 'Confirmed',
            'cancelled': 'Cancelled'
        }
        attendees.append({
            'name': f"{reg.user.first_name} {reg.user.last_name}".strip() or reg.user.username,
            'email': reg.user.email,
            'event_title': reg.event.title,
            'status': status_mapping.get(reg.status, reg.status),
            'rsvp_date': reg.registration_date.date(),
            'registration': reg  # for actions if needed
        })

    # Calculate stats
    confirmed_count = sum(1 for a in attendees if a['status'] == 'Confirmed')
    pending_count = sum(1 for a in attendees if a['status'] == 'Pending')
    cancelled_count = sum(1 for a in attendees if a['status'] == 'Cancelled')

    # Get all emails for mailto
    all_emails = ','.join([a['email'] for a in attendees])

    context = {
        'attendees': attendees,
        'confirmed_count': confirmed_count,
        'pending_count': pending_count,
        'cancelled_count': cancelled_count,
        'all_emails': all_emails,
    }

    return render(request, "localcommunity/eventstudio/attendeesevent.html", context)


# def analyticsevent(request):

#     # ---------------- DASHBOARD COUNTS ----------------
#     total_rsvps = EventRegistration.objects.filter(status='registered').count()

#     total_events = EventRegistration.objects.values('event').distinct().count()

#     attended = EventRegistration.objects.filter(status='attended').count()

#     avg_attendance = (attended / total_rsvps * 100) if total_rsvps > 0 else 0


#     # ---------------- MONTHLY CHART ----------------
#     data = EventRegistration.objects.annotate(
#         month=TruncMonth('registration_date')
#     ).values('month').annotate(
#         total=Count('id')
#     ).order_by('month')

#     months = []
#     counts = []

#     for d in data:
#         months.append(d['month'].strftime('%b'))  # Jan, Feb
#         counts.append(d['total'])


#     # ---------------- EVENT PERFORMANCE ----------------
#     from events.models import Event

#     events = Event.objects.annotate(
#         rsvp_count=Count('eventregistration')
#     )


#     return render(request, 'localcommunity/eventstudio/analyticsbusiness.html', {
#         'total_rsvps': total_rsvps,
#         'total_events': total_events,
#         'avg_attendance': round(avg_attendance, 2),

#         'months': months,
#         'counts': counts,
#         'events': events
#     })


@login_required
@role_required(allowed_roles=["event_organizer"])
def settingsevent(request):
    
    user = request.user

    if request.method == 'POST':
        form = EventProfileUpdateForm(request.POST, instance=user)

        if form.is_valid():

            user_obj = form.save(commit=False)

            current_password = form.cleaned_data.get('current_password')
            new_password = form.cleaned_data.get('new_password')
            confirm_password = form.cleaned_data.get('confirm_password')

            if current_password or new_password or confirm_password:

                if not user.check_password(current_password):
                    messages.error(request, "Current password is incorrect")
                    return render(request, 'localcommunity/eventstudio/settingsevent.html', {"form": form})

                if new_password != confirm_password:
                    messages.error(request, "New passwords do not match")
                    return render(request, 'localcommunity/eventstudio/settingsevent.html', {"form": form})

                user_obj.set_password(new_password)
                user_obj.save()
                update_session_auth_hash(request, user_obj)

                messages.success(request, "Password updated successfully")

            else:
                user_obj.save()
                messages.success(request, "Profile updated successfully")

            return redirect('localcommunity:eventstudiodashboard')

        else:       
            print(form.errors)

    else:
        form = EventProfileUpdateForm(instance=user)
    
    # ✅ FIX IS HERE
    return render(request, "localcommunity/eventstudio/settingsevent.html", {"form": form})


def logoutevent(request):
    logout(request)
    return redirect('login')  # change if needed